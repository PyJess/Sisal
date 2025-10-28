import json
import os
import pandas as pd
from collections import defaultdict
import subprocess
import re
from pathlib import Path
from typing import Dict, List, Tuple, Any
from llm.llm import a_invoke_model
from langchain_openai import ChatOpenAI
from openpyxl import load_workbook
from openpyxl.styles import Font

def load_file(filepath:str):
    with open(filepath, encoding="utf-8") as f:
        return f.read()


def load_json(filepath:str):
    with open(filepath, encoding="utf-8") as f:
        return json.load(f)


def save_json(data, path):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


def get_user_path(user_id: str, subfolder: str = "") -> str:
    """Get the user-specific path."""
    base_dir = os.path.join(os.path.dirname(__file__),"outputs", user_id, subfolder)
    os.makedirs(base_dir, exist_ok=True)
    return base_dir


def excel_to_json(path):
    df = pd.read_excel(path)
    df = df.dropna(axis=1, how="all")  
    df = df.ffill()     
    df.columns = [str(c).strip() for c in df.columns]  

    step_cols = [c for c in df.columns if "step" in c.lower() or "result" in c.lower()]
    id_col = next((c for c in df.columns if c.lower() == "id"), None)

    if not id_col:
        raise ValueError("❌ Nessuna colonna 'ID' trovata nel file Excel!")

    # Tutte le altre colonne tranne gli step
    meta_cols = [c for c in df.columns if c not in step_cols]

    tests = {}

    for _, row in df.iterrows():
        test_id = str(row[id_col]).strip()
        if test_id not in tests:
            meta_data = {col: row.get(col, "") for col in meta_cols if col != id_col}
            meta_data["Steps"] = []
            tests[test_id] = meta_data

        step_data = {col: row.get(col, "") for col in step_cols}
        tests[test_id]["Steps"].append(step_data)

    output_dir = os.path.join(os.path.dirname(__file__),"..", "input")
    output_path = os.path.join(output_dir, "tests_output.json")

    # Scrive il file JSON
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(tests, f, indent=2, ensure_ascii=False)

    print(f"✅ JSON creato in: {output_path}")
    return tests



def group_by_funzionalita(data):
    grouped = defaultdict(dict)

    for key, value in data.items():
        funzionalita = value.get("Funzionalità", "Unknown")
        grouped[funzionalita][key] = value

    # Print nicely
    #print(json.dumps(grouped, indent=2, ensure_ascii=False))
    return (json.dumps(grouped, indent=2, ensure_ascii=False))

PANDOC_EXE = "pandoc"
def process_docx(docx_path, output_base):
    """
    Process a DOCX file using Pandoc and split it into sections based on Markdown headers (#, ##, etc.).
    """
   
    txt_output_path = os.path.join(output_base, Path(docx_path).stem + ".txt")
   
    docx_path = os.path.normpath(docx_path)
    txt_output_path = os.path.normpath(txt_output_path)
    os.makedirs(output_base, exist_ok=True)
   
    # Convert in md
    command = [
        PANDOC_EXE,
        "-s", docx_path,
        "--columns=120",
        "-t", "markdown",
        "-o", txt_output_path
    ]
   
    try:
        subprocess.run(command, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=30)
    except Exception as e:
        print(f"[ERROR] Pandoc conversion failed: {e}")
        return [], os.path.basename(docx_path), []
   
    with open(txt_output_path, "r", encoding="utf-8") as f:
        text_lines = f.read().splitlines()
   
    headers = []
    heading_list = []
   
    for index, line in enumerate(text_lines):
        if line.startswith("#"):
            level = line.count("#")
            clean_name = line.replace("#", "").strip()
            headers.append([clean_name, index, level])
            heading_list.append([clean_name, level])
   
   
    headers.insert(0, ["== first line ==", 0, 0])
    headers.append(["== last line ==", len(text_lines), 0])
    heading_list.insert(0, ["== first line ==", 0])
    heading_list.append(["== last line ==", 0])
   
    head=[]
    chunks = []
    for i in range(len(headers) - 1):
        start_idx = headers[i][1]
        end_idx = headers[i + 1][1]
        section_lines = text_lines[start_idx:end_idx]
        chunk_text = "\n".join(section_lines).strip()
       
        header_cleaned = re.sub(r"\s*\{.*?\}", "", headers[i][0])
        header_cleaned = header_cleaned.replace("--", "–").strip(" *[]\n")
        chunk_text = header_cleaned + "\n" + chunk_text
 
        head.append(header_cleaned)
        chunks.append(chunk_text)
        #print(f" Paragrafo {i+1}: {chunk_text} \n\n")
        print(f" paragrafo {i+1} {header_cleaned}")
   
    return chunks, head




from openpyxl.styles import Font
import pandas as pd
from openpyxl import load_workbook
import os, re

def fill_excel_file(test_cases: dict, output_path: str = None):
    """
    Salva i test case in un file Excel, mantenendo gli step su righe separate.
    Evidenzia in rosso i test generati dall'AI (marcati con [[RED]]...[[/RED]]).
    """
    field_mapping = {
        'Canale': 'Channel',
        'Dispositivo': 'Device',
        'Sistema di riferimento': 'Reference System',
        'Modalità Operativa': 'Execution Mode',
        'Funzionalità': 'Functionality',
        'Tipologia Test': 'Test Type',
        'Test di no regression': 'No Regression Test',
        'Automation': 'Automation',
        'Risultato Atteso': 'Expected Result',
        '_polarion': '_polarion'
    }

    columns = [
        'Title', 'ID', '#', 'Test Group', 'Channel', 'Device', 
        'Priority', 'Test Stage', 'Reference System', 
        'Preconditions', 'Execution Mode', 'Functionality', 
        'Test Type', 'No Regression Test', 'Automation',
        'Dataset', 'Expected Result', 
        'Step', 'Step Description', 'Step Expected Result',
        'Country', 'Project', 'Author', 'Assignee(s)', 'Type', 
        'Partial Coverage Description', '_polarion',
        'Analysis', 'Coverage', 'Dev Complexity', 'Execution Time', 
        'Volatility', 'Developed', 'Note', 'Team Ownership', 
        'Team Ownership Note', 'Requires Script Maintenance'
    ]

    rows = []
    for tc_id, tc_data in test_cases.items():
        steps = tc_data.get('Steps', [])
        if not steps:
            steps = [{}]
        
        first = True
        for step in steps:
            row = {}
            if first:
                for col in columns:
                    if col not in ['Step', 'Step Description', 'Step Expected Result']:
                        value = tc_data.get(col, '')
                        if not value:
                            italian_key = next((k for k, v in field_mapping.items() if v == col), None)
                            if italian_key:
                                value = tc_data.get(italian_key, '')
                        row[col] = value
                first = False
            else:
                for col in columns:
                    if col not in ['Step', 'Step Description', 'Step Expected Result']:
                        row[col] = ''

            row['Step'] = step.get('Step', '')
            row['Step Description'] = step.get('Step Description', '')
            row['Step Expected Result'] = step.get('Expected Result', '')
            rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    if output_path is None:
        output_path = os.path.join(os.path.dirname(__file__), "..", "outputs", "testbook_feedbackAI.xlsx")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    df.to_excel(output_path, index=False)

    wb = load_workbook(output_path)
    ws = wb.active

    red_font = Font(color="FF0000")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "[[RED]]" in cell.value:
                clean_text = re.sub(r'\[\[/?RED\]\]', '', cell.value)
                cell.value = clean_text
                cell.font = red_font  # 🔴 colora tutto il testo della cella in rosso

    wb.save(output_path)
    print(f"✅ Excel salvato con testi rossi: {output_path}")

def prepare_test_texts(df):
    """
    Combina i campi di ogni test case in un unico testo da usare per embedding o LLM.
    df: DataFrame Pandas con colonne tipo Title, Step, Step Description, Expected Result, Preconditions, Funzionalità, Device.
    """
    test_texts = []
    
    for idx, row in df.iterrows():
        # Costruisci testo degli step
        step_num = row.get("Step", "")
        step_desc = row.get("Step Description", "")
        expected = row.get("Expected Result", "")
        steps_text = f"Step {step_num}: {step_desc}. Expected: {expected}. "
        
        # Combina tutto in un unico testo
        combined_text = (
            f"Title: {row.get('Title','')}. "
            f"Functionality: {row.get('Funzionalità','')}. "
            f"Preconditions: {row.get('Preconditions','')}. "
            f"Steps: {steps_text}"
        )
        
        test_texts.append({
            "id": row.get("ID", idx),
            "text": combined_text
        })
    
    return test_texts


def prepare_prompt_requisiti(req: str,context: str, mapping: str = None) -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    """Prepare prompt for the LLM"""
    system_prompt = load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "copertura_requisiti", "system_prompt.txt"))
    user_prompt = load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "copertura_requisiti", "user_prompt.txt")) 
    schema = load_json(os.path.join(os.path.dirname(__file__), "..", "llm", "schema", "schema_output.json"))

    user_prompt = user_prompt.replace("{req}", req)
    user_prompt = user_prompt.replace("{context}", context)
    mapping_as_string = mapping.to_json() 
    user_prompt = user_prompt.replace("{mapping}", mapping_as_string)

    print("finishing prepare prompt")

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    return messages, schema


def a_invoke_model_requisiti(msgs, schema, model="gpt-4.1"):
        """Invoke the LLM model"""

        gpt = ChatOpenAI(model=model, temperature=0.1).with_structured_output(schema=schema, strict=True)


        result =  gpt.invoke(msgs)
    
        # Estrai i token usage
        if isinstance(result, dict) and 'raw' in result:
            usage = result['raw'].response_metadata.get('token_usage', {})
            print(f"\n📊 Token Usage:")
            print(f"  Input tokens:  {usage.get('prompt_tokens', 0)}")
            print(f"  Output tokens: {usage.get('completion_tokens', 0)}")
            print(f"  Total tokens:  {usage.get('total_tokens', 0)}")
            
            # Restituisci solo i dati parsati
            return result['parsed']
        
        return result
        #return await gpt.ainvoke(msgs)


def AI_check_TC_requisiti(req: str,context: str, mapping: str = None) -> Dict:
    messages, schema =  prepare_prompt_requisiti(req,context, mapping)
    print("starting calling llm")
    print(f"{messages}")
    response = a_invoke_model_requisiti(messages, schema, model="gpt-4.1")
    return response


def color_new_testcases_red(excel_path: Path, new_rows_count: int):
    """
    Colora di rosso le ultime `new_rows_count` righe nel file Excel.
    """
    wb = load_workbook(excel_path)
    ws = wb.active

    # Stili di font rosso
    red_font = Font(color="FF0000")  # rosso acceso

    max_row = ws.max_row
    max_col = ws.max_column

    # Righe da colorare: ultime `new_rows_count`
    start_row = max_row - new_rows_count + 1

    for row in ws.iter_rows(min_row=start_row, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.font = red_font

    wb.save(excel_path)
    wb.close()
    print(f"🟥 Colorate di rosso {new_rows_count} righe in {excel_path.name}")


def fill_excel_file_requisiti(test_cases: dict, base_columns=None):
    """
    Converte i test case generati dall'LLM in un DataFrame Excel compatibile
    con la struttura del file di input.
    Ogni step viene riportato su una riga separata.
    """

    # 🔹 Colonne di default (in caso non vengano passate)
    if base_columns is None:
        base_columns = [
            'Title', 'ID', '#', 'Test Group', 'Channel', 'Device',
            'Priority', 'Test Stage', 'Reference System',
            'Preconditions', 'Execution Mode', 'Functionality',
            'Test Type', 'No Regression Test', 'Automation',
            'Dataset', 'Expected Result',
            'Step', 'Step Description', 'Step Expected Result',
            'Country', 'Project', 'Author', 'Assignee(s)', 'Type',
            'Partial Coverage Description', '_polarion',
            'Analysis', 'Coverage', 'Dev Complexity', 'Execution Time',
            'Volatility', 'Developed', 'Note', 'Team Ownership',
            'Team Ownership Note', 'Requires Script Maintenance'
        ]

    # Mappa per eventuali chiavi italiane
    field_mapping = {
        'Canale': 'Channel',
        'Dispositivo': 'Device',
        'Sistema di riferimento': 'Reference System',
        'Modalità Operativa': 'Execution Mode',
        'Funzionalità': 'Functionality',
        'Tipologia Test': 'Test Type',
        'Test di no regression': 'No Regression Test',
        'Automation': 'Automation',
        'Risultato Atteso': 'Expected Result',
        '_polarion': '_polarion'
    }

    rows = []

    # Gestione dei vari livelli annidati
    for tc_group in test_cases.values():
        if isinstance(tc_group, dict) and "test_cases" in tc_group:
            tcs = tc_group["test_cases"]
        elif isinstance(tc_group, list):
            tcs = tc_group
        else:
            tcs = [tc_group]

        for tc_data in tcs:

            steps = tc_data.get("Steps", [])
            if not steps:
                steps = [{}]

            for i, step in enumerate(steps):
                row = {}

                # Prima riga → tutti i dati generali del test case
                if i == 0:
                    for col in base_columns:
                        if col not in ['Step', 'Step Description', 'Step Expected Result']:
                            value = tc_data.get(col, '')
                            if not value:
                                italian_key = next((k for k, v in field_mapping.items() if v == col), None)
                                if italian_key:
                                    value = tc_data.get(italian_key, '')
                            row[col] = value
                else:
                    for col in base_columns:
                        if col not in ['Step', 'Step Description', 'Step Expected Result']:
                            row[col] = ''

                # Inserisci i dati dello step
                row['Step'] = step.get('Step', '')
                row['Step Description'] = step.get('Step Description', '')
                row['Step Expected Result'] = step.get('Expected Result', '')

                rows.append(row)

    df = pd.DataFrame(rows, columns=base_columns)
    df = df.loc[:, ~df.columns.duplicated()]

    print(f"Generato DataFrame con {len(df)} righe e {len(df.columns)} colonne")
    print(f"Prime colonne: {list(df.columns[:6])}")
    print(f"Esempio step: {df[['Title','Step','Step Description']].head(3)}")

    return df
