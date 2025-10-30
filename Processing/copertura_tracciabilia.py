import sys
import os
from pathlib import Path
from openpyxl import Workbook
import json
from docx import Document
import subprocess
from langchain_openai import OpenAIEmbeddings
from langchain.vectorstores import FAISS
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import asyncio

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from utils.simple_functions import group_by_funzionalita, load_json
from Processing.controllo_sintattico import *
import json
from Processing.test_design import create_vectordb

embedding_model = "text-embedding-3-large"

async def prepare_prompt(input: str, context:str ="", testcase: Dict = None) -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    """Prepare prompt for the LLM"""
    system_prompt = load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "copertura_tracciabilita", "system_prompt.txt"))
    user_prompt = load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "copertura_tracciabilita", "user_prompt.txt")) 
    schema = load_json(os.path.join(os.path.dirname(__file__), "..", "llm", "schema", "schema_output_tracciabilita.json"))

    user_prompt = user_prompt.replace("{input}", input)
    print(f"input {input}")
    user_prompt = user_prompt.replace("{context}", context )
    testcase= json.dumps(testcase, ensure_ascii=False, indent=2)
    user_prompt = user_prompt.replace("{testcase}", testcase )
    print("finishing prepare prompt")

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    return messages, schema


async def AI_gen_title(input: str, context:str, testcase: Dict = None) -> Dict:
    messages, schema = await prepare_prompt(input, context, testcase)
    print("starting calling llm")
    #print(f"{messages}")
    response = await a_invoke_model(messages, schema, model="gpt-4.1")
    return response



async def main():
    excel_path = os.path.join(os.path.dirname(__file__), "..", "outputs", "generated_test_cases3_withoutDHW.xlsx")
    dic = excel_to_json(excel_path) 
    print("finishing excel to json")

    rag_path = os.path.join(os.path.dirname(__file__), "..", "input", "Esempio 2", 
                            "RU_Sportsbook_Platform_Fantacalcio_Prob. Form_v0.2 (1).docx")
    chunks, _ = process_docx(rag_path, os.path.dirname(rag_path))
    embeddings = OpenAIEmbeddings(model=embedding_model)
    vectorstore = FAISS.from_texts(chunks, embeddings)

    for test_id, testcase in dic.items():
        Title = testcase['Title']
        tc=str(testcase)
        context = create_vectordb(tc, vectorstore, k=3, similarity_threshold=0.75)
        context=""
        new_Title = await AI_gen_title(Title, context, testcase)
        testcase['Title'] = new_Title["corrected_title"]
        print(f"Corrected Title for {test_id}: {new_Title}")

    
    excel_path = os.path.join(os.path.dirname(__file__), "..", "outputs", "testbook_tracciabilita_feedbackAI.xlsx")
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    df1=convert_to_DF(dic)
    df1.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    wb.save(excel_path)

if __name__ == "__main__":
    asyncio.run(main())