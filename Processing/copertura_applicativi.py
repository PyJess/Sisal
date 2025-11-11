import sys
import os
from pathlib import Path
from openpyxl import Workbook
import json
from docx import Document
import subprocess
from langchain_openai import OpenAIEmbeddings
from langchain.vectorstores import FAISS
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import PatternFill

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from utils.simple_functions import group_by_funzionalita, load_json
from Processing.controllo_sintattico import *
import json
from Processing.test_design import create_vectordb
from llm.llm import LLMClient
from dotenv import load_dotenv
load_dotenv()


llm_client = LLMClient()

#embedding_model = "text-embedding-3-large"
PANDOC_EXE = "pandoc" 



def convert_to_DF(test_cases: dict):
    """
    Structure the test cases into a DataFrame suitable for Excel export,
    """
    field_mapping = {
        'Canale': 'Channel',
        'Dispositivo': 'Device',
        'Sistema di riferimento': 'Reference System',
        'Modalità Operativa': 'Execution Mode',
        'Funzionalità': 'Functionality',
        'Tipologia Test': 'Test Type',
        'Test di no regression': 'No Regression Test',
        'Automation': 'Automation',
        'Risultato Atteso': 'Expected Result',
        '_polarion': '_polarion'
    }
    
    # Definisci colonne finali (solo inglese)
    columns = [
        'Title', 'ID', '#', 'Test Group', 'Channel', 'Device', 
        'Priority', 'Test Stage', 'Reference System', 
        'Preconditions', 'Execution Mode', 'Functionality', 
        'Test Type', 'No Regression Test', 'Automation',
        'Dataset', 'Expected Result', 
        'Step', 'Step Description', 'Step Expected Result',
        'Country', 'Project', 'Author', 'Assignee(s)', 'Type', 
        'Partial Coverage Description', '_polarion',
        'Analysis', 'Coverage', 'Dev Complexity', 'Execution Time', 
        'Volatility', 'Developed', 'Note', 'Team Ownership', 
        'Team Ownership Note', 'Requires Script Maintenance'
    ]

    rows = []
    for tc_id, tc_data in test_cases.items():
        steps = tc_data.get('Steps', [])
        
        if not steps:
            steps = [{}]
        
        first = True
        for step in steps:
            row = {}

            if first:

                for col in columns:
                    if col not in ['Step', 'Step Description', 'Step Expected Result']:
                        value = tc_data.get(col, '')
                        
                        if not value:
                            italian_key = next((k for k, v in field_mapping.items() if v == col), None)
                            if italian_key:
                                value = tc_data.get(italian_key, '')
                        
                        row[col] = value
                first = False
            else:
                for col in columns:
                    if col not in ['Step', 'Step Description', 'Step Expected Result']:
                        row[col] = ''

            row['Step'] = step.get('Step', '')
            row['Step Description'] = step.get('Step Description', '')
            row['Step Expected Result'] = step.get('Expected Result', '')
            
            rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    return df


async def prepare_prompt_application(input: str, mapping: str = None) -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    """Prepare prompt for the LLM"""
    system_prompt = load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "copertura_applicativi", "system_prompt.txt"))
    user_prompt = load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "copertura_applicativi", "user_prompt.txt")) 
    schema = load_json(os.path.join(os.path.dirname(__file__), "..", "llm", "schema", "schema_output_applicativi.json"))

    user_prompt = user_prompt.replace("{input}", input)

    print("finishing prepare prompt")

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    return messages, schema


async def AI_check_applications(input: str, mapping: str = None) -> Dict:
    messages, schema = await prepare_prompt_application(input)
    print("starting calling llm")
    #print(f"{messages}")
    response = await llm_client.a_invoke_model(messages, schema)
    return response


async def prepare_prompt(input: str, context:str = None, mapping: str = None) -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    """Prepare prompt for the LLM"""
    system_prompt = load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "test_design", "system_prompt.txt"))
    user_prompt = load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "test_design", "user_prompt.txt")) 
    schema = load_json(os.path.join(os.path.dirname(__file__), "..", "llm", "schema", "schema_output.json"))

    user_prompt = user_prompt.replace("{input}", input)
    print(f"input {input}")
    user_prompt = user_prompt.replace("{context}", context )
    mapping= mapping.to_json() 
    user_prompt = user_prompt.replace("{mapping}", mapping )
    print("finishing prepare prompt")

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    return messages, schema


async def AI_gen_TC(input: str, context:str, mapping: str = None) -> Dict:
    messages, schema = await prepare_prompt(input, context, mapping)
    print("starting calling llm")
    #print(f"{messages}")
    response = await llm_client.a_invoke_model(messages, schema)
    return response

async def main():
    excel_path = os.path.join(os.path.dirname(__file__), "..", "outputs", "generated_test_feedbackAI.xlsx")
    dic = excel_to_json(excel_path) 
    print("finishing excel to json")

    mapping = extract_field_mapping()
    
    application_list = []
    input_path=os.path.join(os.path.dirname(__file__), "..", "input", "Esempio 2", "RU_Sportsbook_Platform_Fantacalcio_Prob. Form_v0.2 (1).docx")
    requirements, title = process_docx(input_path, os.path.dirname(input_path))
    
    rag_path= input_path=os.path.join(os.path.dirname(__file__), "..", "input", "Esempio 2", "RU_Sportsbook_Platform_Fantacalcio_Prob. Form_v0.2 (1).docx")
    chunks, _ = process_docx(input_path, os.path.dirname(rag_path))
    #embeddings = OpenAIEmbeddings(model=embedding_model)
    chunks= chunks + requirements
    #vectorstore = FAISS.from_texts(chunks, embeddings)
    
    tasks = [AI_check_applications(input=req) for req in requirements]
    applications_results = await asyncio.gather(*tasks)

    normalized_results = []
    for res in applications_results:
        if isinstance(res, dict) and "applications" in res:
            normalized_results.append(res["applications"])
        elif isinstance(res, list):  
            normalized_results.append(res)
        else:
            normalized_results.append([])

    application_list = []
    for t, apps in zip(title, normalized_results):
        application_list.append({
            "title": t,
            "applications": apps
        })

    # Eliminates duplicates based on application_name
    unique_apps_dict = {}

    for sublist in application_list:
        for app in sublist.get("applications", []):
            name = app.get("application_name")
            if name:
                unique_apps_dict[name.strip()] = {
                    "title": sublist.get("title"),
                    "application_name": name.strip(),
                    "specific_text": app.get("specific_text", "").strip()
                }

    unique_apps = list(unique_apps_dict.values())

    application_list = [{"applications": unique_apps}]

    print("Applications found: \n\n")
    for app in application_list[0]["applications"]:
        print(app.get("application_name"))


    for result in application_list:
        new_TC=[]
        for app in result.get("applications", []):
            app_name = app.get("application_name", "").strip()
            app_title= app.get("title", "").strip()
            print(f"\n{'='*50}")
            print(f"Checking application: '{app_name}'")
            print(app_title)
            present = False
            
            test_cases = dic.values() 
            
            print(f"Total test cases to check: {len(list(dic.values()))}")
            
            for tc_id, test_case in dic.items(): 
                if not isinstance(test_case, dict):
                    print(f"  {tc_id}: Not a dict, skipping")
                    continue
                
                # Controlla Title
                title = test_case.get("Title", "")
                
                if app_name.lower() in title.lower():
                    print(f"  {tc_id}: FOUND in title!")
                    present = True
                    break
                
                # Controlla tutti gli step
                steps = test_case.get("Steps") or test_case.get("Step")
                if steps:
                    for j, step in enumerate(steps):
                        for field in ["Step Description", "Expected Result"]:
                            value = step.get(field, "")
                            if app_name.lower() in str(value).lower():
                                print(f"  {tc_id}, Step {j}: FOUND in {field}!")
                                present = True
                                break
                        if present:
                            break
                if present:
                    break
            
            if present:
                print(f"{app_name}: presente")
            else:
                print(f"{app_name}: non presente")
                app_text= app.get("specific_text", "")
                vectorstore= None
                #context = create_vectordb(app_text, vectorstore, k=3, similarity_threshold=0.75)
                #context= str(context)
                context=""
                generated_tc= await AI_gen_TC(app_text, context, mapping)
                if "test_cases" in generated_tc:
                    for tc in generated_tc["test_cases"]:
                        tc["_polarion"] = app_title

                new_TC.append(generated_tc)

                while True:
                    generated_tc= await AI_gen_TC(app_text, context, mapping)
                    if isinstance(generated_tc, dict):
                        if "test_cases" in generated_tc:
                            for tc in generated_tc["test_cases"]:
                                tc["_polarion"] = app_title

                        new_TC.append(generated_tc)
                        break
                    print(f"generated tc non valido per '{app_title}', riprovo...")

    all_generated = []

    for result in new_TC:
        all_generated.extend(result.get("test_cases", []))

    new_TC_dict = {tc["ID"]: tc for tc in all_generated}
    df2= convert_to_DF(dic)
    df2["#"] = pd.to_numeric(df2["#"], errors="coerce")
    max_num = int(df2["#"].max()) if not df2["#"].empty else 0

    for i, tc_id in enumerate(new_TC_dict, start=1):
        new_TC_dict[tc_id]["#"] = max_num + i

    excel_path = os.path.join(os.path.dirname(__file__), "..", "outputs", "testbook_applicativi_feedbackAI.xlsx")
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    df1=convert_to_DF(new_TC_dict)

    df_combined = pd.concat([df2, df1], ignore_index=True)
    df_combined.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    start_row = 2 + len(df2)  
    end_row = start_row + len(df1) 
    for row_idx in range(start_row, end_row):
        for cell in ws[row_idx]:
            if cell.value is not None:
                cell.font = Font(color="FF0000")


    wb.save(excel_path)
    

if __name__ == "__main__":
    asyncio.run(main())

