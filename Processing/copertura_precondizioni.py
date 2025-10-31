import pandas as pd
from pathlib import Path
from langchain_openai import OpenAIEmbeddings
from langchain_community.vectorstores import FAISS
import sys
import os
from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from openpyxl import load_workbook
from openpyxl.styles import Font
from langchain_openai import ChatOpenAI

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from utils.simple_functions import process_docx,load_file

load_dotenv()

def agent_preconditions(data_sample: str, head: str, chunks: str, model="gpt-4.1"):

    gpt = ChatOpenAI(model=model, temperature=0.1)

    #Percorso ai prompt
    base_path = os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "copertura_precondizioni")

    system_prompt = load_file(os.path.join(base_path, "system_prompt.txt"))
    user_prompt = load_file(os.path.join(base_path, "user_prompt.txt"))
    
    #Sostituzione dei placeholder nel prompt utente
    user_prompt = user_prompt.replace("{head}", str(head))
    user_prompt = user_prompt.replace("{chunks}", str(chunks))
    user_prompt = user_prompt.replace("{data_sample}", str(data_sample))

  
    messages = [
        {"role": "system", "content": system_prompt.strip()},
        {"role": "user", "content": user_prompt.strip()},
    ]

    response = gpt.invoke(messages)
    return response.content

input_excel = Path(__file__).parent.parent/"input"/"generated_test_cases3 - Copia.xlsx"
# testcase precondizione mancante = "SGP Generator application is deployed and accessible via a supported desktop browser."
# User is not logged in or is logged in; Home Page is accessible; SGP cards are available from BE engine.
output_excel = input_excel.with_name(f"{input_excel.stem}_feedbackAI_precondizioni.xlsx")

embeddings = OpenAIEmbeddings(model="text-embedding-3-large")

path_document_word = Path(__file__).parent.parent/"input"/"RU_SportsBookPlatform_SGP_Gen_FullResponsive_v1.1 - FE (2).docx"

path_output = Path(__file__).parent.parent/"outputs"

chunks,head= process_docx(path_document_word,path_output)

df = pd.read_excel(input_excel)

        
# Filtra solo le righe che hanno un titolo per skippare le celle che sono step e non testcase
df_cases = df[df["Title"].notna() & (df["Title"].astype(str).str.strip() != "")]
print(f"Trovati {len(df_cases)} test case principali su {len(df)} righe totali")



for idx, case in df_cases.iterrows():
    title = str(case.get("Title", "")).strip()
    precond = case.get("Preconditions", "")
    polarion = str(case.get("_polarion", "")).lower().strip()

    context = ""

    #Controlla se la precondizione manca
    if pd.isna(precond) or str(precond).strip() == "":
        print(f"⚠️ Manca precondizione per: {title}")

        #Trova il paragrafo corrispondente nel documento Word
        for req in head:
            req_norm = req.lower().strip()
            if req_norm == polarion:
                context = chunks[head.index(req)]
                # print(req)
                # print("*********")
                # print(context)
                break

        #Fallback se nessuna sezione trovata
        if not context:
            context = "No matching documentation section found."

        #Ricostruisco il data sample
        data_sample = "\n".join(
            [f"{col}: {val}" for col, val in case.items() if pd.notna(val) and str(val).strip() != ""]
        )

        #Chiamata all’agente AI per generare la precondizione
        generated_precond = agent_preconditions( data_sample,req,context)

        precond_text = f"[red] {generated_precond.strip()}"


        df.at[idx, "Preconditions"] = precond_text

    else:
        print(f"{title}: precondizione presente -> {precond}")
   


df.to_excel(output_excel, index=False)
print(f"File aggiornato salvato in: {output_excel}")



wb = load_workbook(output_excel)
ws = wb.active

col_index = df.columns.get_loc("Preconditions") + 1

for row in range(2, ws.max_row + 1):  
    cell = ws.cell(row=row, column=col_index)
    if cell.value and "[red]" in str(cell.value):
        cell.font = Font(color="FF0000")
        cell.value = cell.value.replace("[red]", "").strip()

wb.save(output_excel)
print(" Celle con precondizioni AI colorate di rosso.")