import os
import sys
import pandas as pd
# Ensure project root is on sys.path so local packages can be imported
# when this module is executed directly (python Processing/controllo_sintattico.py)
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from llm.llm import a_invoke_model
from utils.simple_functions import *
import asyncio
from typing import Tuple, List, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
import re
from Input_extraction.extract_polarion_field_mapping import *
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock


async def prepare_prompt(input:Dict, mapping:str =None) -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    """ Prepare prompt for the LLM"""

    system_prompt= load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "controllo_sintattico", "system_prompt.txt"))
    user_prompt= load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "controllo_sintattico", "user_prompt.txt")) 
    schema= load_json(os.path.join(os.path.dirname(__file__), "..", "llm", "schema", "schema_output.json"))

    user_prompt=user_prompt.replace("{input}", json.dumps(input))
    mapping_as_string = mapping.to_json() 
    user_prompt = user_prompt.replace("{mapping}", mapping_as_string)
    print("finishing prepare prompt")

    messages=[
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    return messages, schema


async def AI_check_TC(input:Dict, mapping:str =None) -> Dict:

    #input = json.loads(input)
    messages, schema= await prepare_prompt(input, mapping)
    print("starting calling llm")
    print(f"{messages}")
    response = await a_invoke_model(messages, schema)

    return response


def apply_red_text(cell):
        """Color text in [[RED]]...[[/RED]] red, preserving the rest."""

        text = str(cell.value)
        if "[[RED]]" not in text:
            return  

        parts = re.split(r'(\[\[RED\]\]|\[\[/RED\]\])', text)
        rich_text = CellRichText()

        red = False
        for part in parts:
            if part == "[[RED]]":
                red = True
            elif part == "[[/RED]]":
                red = False
            elif part:
                rich_text.append(TextBlock(InlineFont(color="FF0000") if red else InlineFont(color="000000"), part))

        cell.value = rich_text



def fill_excel_file(test_cases: dict):
    """
    Salva i test case in Excel, mantenendo gli step su righe separate
    e applica i testi rossi dove necessario.
    """
    # Costruzione righe per DataFrame
    columns = [
        'Title', 'ID', 'Test Group', 'Channel', 'Device', 'Priority',
        'Test Stage', 'Reference System', 'Preconditions', 'Execution Mode',
        'Functionality', 'Test Type', 'No Regression Test', 'Automation',
        'Expected Result', 'Step', 'Step Description', '_polarion'
    ]

    rows = []
    for tc_id, tc_data in test_cases.items():
        steps = tc_data.get('Steps', [])
        first = True
        for step in steps:
            row = {}
            if first:
                for col in columns[:15]:
                    row[col] = tc_data.get(col, '')
                first = False
            else:
                for col in columns[:15]:
                    row[col] = ''
            row['Step'] = step.get('Step', '')
            row['Step Description'] = step.get('Step Description', '')
            row['Step Expected Result'] = step.get('Expected Result', '')
            rows.append(row)

    df = pd.DataFrame(rows, columns=columns)

    # Percorso Excel
    excel_path = os.path.join(os.path.dirname(__file__), "..", "outputs", "testbook_feedbackAI.xlsx")
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    # 1️⃣ Prima: salva il file Excel (senza formattazione)
    df.to_excel(excel_path, index=False)

    # 2️⃣ Poi: riapri con openpyxl e applica apply_red_text
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "[[RED]]" in cell.value:
                apply_red_text(cell)

    # 3️⃣ Salva di nuovo con il testo colorato
    wb.save(excel_path)
    print(f"✅ Excel salvato con testi rossi: {excel_path}")


async def main():
    mapping = extract_field_mapping()
    print("finishing mapping")
    input_path = os.path.join(os.path.dirname(__file__), "..", "input", "tests_cases.xlsx")
    dic = excel_to_json(input_path) 
    print("finishing excel to json")
 
    tasks = [AI_check_TC(input={"ID": tc_id, **tc_data}, mapping=mapping) for tc_id, tc_data in dic.items()]
    results_list = await asyncio.gather(*tasks)
    print("finishing gpt call")
    
    merged_results = {tc["ID"]: tc for tc in results_list}


    print(f"Merged result: {merged_results}")

    fill_excel_file(merged_results)


if __name__ == "__main__":
    asyncio.run(main())

