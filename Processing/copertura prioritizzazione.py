import pandas as pd
from pathlib import Path
from langchain_openai import OpenAIEmbeddings
from langchain_community.vectorstores import FAISS
import sys
import os
from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from openpyxl import load_workbook
from openpyxl.styles import Font
from langchain_openai import ChatOpenAI

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from utils.simple_functions import process_docx,load_file

load_dotenv()

def agent_prioritizzazione(data_sample: str, head: str, chunks: str, model="gpt-4.1"):

    gpt = ChatOpenAI(model=model, temperature=0.1)

    #Percorso ai prompt
    base_path = os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "copertura_prioritizzazione")

    system_prompt = load_file(os.path.join(base_path, "system_prompt.txt"))
    user_prompt = load_file(os.path.join(base_path, "user_prompt.txt"))
    
    #Sostituzione dei placeholder nel prompt utente
    user_prompt = user_prompt.replace("{head}", str(head))
    user_prompt = user_prompt.replace("{chunks}", str(chunks))
    user_prompt = user_prompt.replace("{data_sample}", str(data_sample))

  
    messages = [
        {"role": "system", "content": system_prompt.strip()},
        {"role": "user", "content": user_prompt.strip()},
    ]

    response = gpt.invoke(messages)
    return response.content

input_excel = Path(__file__).parent.parent/"input"/"tests_cases.xlsx"


path_output = Path(__file__).parent.parent / "outputs"

output_excel = path_output / f"{input_excel.stem}_feedbackAI_priority.xlsx"

embeddings = OpenAIEmbeddings(model="text-embedding-3-large")

path_document_word = Path(__file__).parent.parent/"input"/"RU_SportsBookPlatform_SGP_Gen_FullResponsive_v1.1 - FE (2).docx"

path_output = Path(__file__).parent.parent/"outputs"

chunks,head= process_docx(path_document_word,path_output)

df = pd.read_excel(input_excel)

# Filtra solo le righe che hanno un titolo per skippare le celle che sono step e non testcase
df_cases = df[df["Title"].notna() & (df["Title"].astype(str).str.strip() != "")]
#df_cases = df_cases.head(20) # per test rapidi 

print(f"Trovati {len(df_cases)} test case principali su {len(df)} righe totali")



for idx, case in df_cases.iterrows():
    title = str(case.get("Title", "")).strip()
    current_priority = str(case.get("Priority", "")).strip()
    polarion = str(case.get("_polarion", "")).lower().strip()
    req = "Unknown requirement"
    context = ""

    # Trova il contesto nel documento Word (stesso meccanismo di prima)
    for req in head:
        req_norm = req.lower().strip()
        if req_norm == polarion:
            context = chunks[head.index(req)]
            # print(req)
            # print("*********")
            # print(context)
            break

    # Ricostruzione del contesto testuale da passare al modello
    data_sample = "\n".join(
        [f"{col}: {val}" for col, val in case.items() if pd.notna(val) and str(val).strip() != ""]
    )

    # Invoca il modello AI
    generated_priority = agent_prioritizzazione(data_sample, req, context).strip()

    if not current_priority:  # Mancante → generata ex novo
        priority_text = f"[red]{generated_priority}"
    elif generated_priority.lower() != current_priority.lower():  # Diversa → modificata
        priority_text = f"[red]{generated_priority}"
    else:  # Uguale → lascia invariata
        priority_text = current_priority

    df.at[idx, "Priority"] = priority_text

df.to_excel(output_excel, index=False)
print(f"File aggiornato salvato in: {output_excel}")


wb = load_workbook(output_excel)
ws = wb.active

col_index = df.columns.get_loc("Priority") + 1

for row in range(2, ws.max_row + 1):  
    cell = ws.cell(row=row, column=col_index)
    if cell.value and "[red]" in str(cell.value):
        cell.font = Font(color="FF0000")
        cell.value = cell.value.replace("[red]", "").strip()

wb.save(output_excel)
print(" Celle con precondizioni AI colorate di rosso.")