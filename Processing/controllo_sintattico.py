import os
import json
import re
import sys
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
import asyncio
from typing import Dict, List, Tuple, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from llm.llm import LLMClient
from utils.simple_functions import *
import asyncio
from typing import Tuple, List, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.cell.rich_text import CellRichText, TextBlock
import re
from Input_extraction.extract_polarion_field_mapping import *
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock

llm_client = LLMClient()

async def prepare_prompt(input: Dict, mapping: str = None) -> Tuple[List[Dict[str, str]], Dict[str, Any]]:
    """Prepare prompt for the LLM"""
    system_prompt = load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "controllo_sintattico", "system_prompt.txt"))
    user_prompt = load_file(os.path.join(os.path.dirname(__file__), "..", "llm", "prompts", "controllo_sintattico", "user_prompt.txt")) 
    schema = load_json(os.path.join(os.path.dirname(__file__), "..", "llm", "schema", "schema_output1.json"))

    user_prompt = user_prompt.replace("{input}", json.dumps(input))
    mapping_as_string = mapping.to_json() 
    user_prompt = user_prompt.replace("{mapping}", mapping_as_string)
    print("finishing prepare prompt")

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": user_prompt},
    ]

    return messages, schema


async def AI_check_TC(input: Dict, mapping: str = None) -> Dict:
    """Call LLM to check test case syntax"""
    messages, schema = await prepare_prompt(input, mapping)
    print("starting calling llm")
    print(f"{messages}")
    response = await llm_client.a_invoke_model(messages, schema)
    #response = await llm_client.a_invoke_model(messages, schema, model="gpt-4o")
    return response


def apply_red_text(cell):
    """Color text in [[RED]]...[[/RED]] red, preserving the rest."""
    text = str(cell.value)
    if "[[RED]]" not in text:
        return  

    parts = re.split(r'(\[\[RED\]\]|\[\[/RED\]\])', text)
    rich_text = CellRichText()

    red = False
    for part in parts:
        if part == "[[RED]]":
            red = True
        elif part == "[[/RED]]":
            red = False
        elif part:
            font = InlineFont(color="FF0000") if red else InlineFont(color="000000")
            rich_text.append(TextBlock(font, part))

    cell.value = rich_text


def fill_excel_file(test_cases: dict):
    """
    Salva i test case in Excel, mantenendo gli step su righe separate
    e applica i testi rossi dove necessario.
    """
    field_mapping = {
        'Canale': 'Channel',
        'Dispositivo': 'Device',
        'Sistema di riferimento': 'Reference System',
        'Modalità Operativa': 'Execution Mode',
        'Funzionalità': 'Functionality',
        'Tipologia Test': 'Test Type',
        'Test di no regression': 'No Regression Test',
        'Automation': 'Automation',
        'Risultato Atteso': 'Expected Result',
        '_polarion': '_polarion'
    }
    
    # Definisci colonne finali (solo inglese)
    columns = [
        'Title', 'ID', '#', 'Test Group', 'Channel', 'Device', 
        'Priority', 'Test Stage', 'Reference System', 
        'Preconditions', 'Execution Mode', 'Functionality', 
        'Test Type', 'No Regression Test', 'Automation',
        'Dataset', 'Expected Result', 
        'Step', 'Step Description', 'Step Expected Result',
        'Country', 'Project', 'Author', 'Assignee(s)', 'Type', 
        'Partial Coverage Description', '_polarion',
        'Analysis', 'Coverage', 'Dev Complexity', 'Execution Time', 
        'Volatility', 'Developed', 'Note', 'Team Ownership', 
        'Team Ownership Note', 'Requires Script Maintenance'
    ]

    rows = []
    for tc_id, tc_data in test_cases.items():
        steps = tc_data.get('Steps', [])
        
        if not steps:
            steps = [{}]
        
        first = True
        for step in steps:
            row = {}

            if first:

                for col in columns:
                    if col not in ['Step', 'Step Description', 'Step Expected Result']:
                        value = tc_data.get(col, '')
                        
                        if not value:
                            italian_key = next((k for k, v in field_mapping.items() if v == col), None)
                            if italian_key:
                                value = tc_data.get(italian_key, '')
                        
                        row[col] = value
                first = False
            else:
                for col in columns:
                    if col not in ['Step', 'Step Description', 'Step Expected Result']:
                        row[col] = ''

            row['Step'] = step.get('Step', '')
            row['Step Description'] = step.get('Step Description', '')
            row['Step Expected Result'] = step.get('Expected Result', '')
            
            rows.append(row)

    df = pd.DataFrame(rows, columns=columns)


    excel_path = os.path.join(os.path.dirname(__file__), "..", "outputs", "testbook_controllosintattico_feedbackAI.xlsx")
    os.makedirs(os.path.dirname(excel_path), exist_ok=True)

    df.to_excel(excel_path, index=False)

    wb = load_workbook(excel_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):  
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "[[RED]]" in cell.value:
                apply_red_text(cell)

    wb.save(excel_path)
    print(f"Excel salvato con testi rossi: {excel_path}")


async def main():
    mapping = extract_field_mapping()
    print("finishing mapping")
    
    input_path = os.path.join(os.path.dirname(__file__), "..", "input", "tests_cases_modified.xlsx")
    dic = excel_to_json(input_path) 
    print("finishing excel to json")
 
    max_attempts = 5
    attempt = 0
    while attempt < max_attempts:
        attempt += 1
        tasks = [AI_check_TC(input={"ID": tc_id, **tc_data}, mapping=mapping) for tc_id, tc_data in dic.items()]
        results_list = await asyncio.gather(*tasks)
        if all(isinstance(res, dict) and res for res in results_list):
            print("finishing gpt call")
            break
        print(f"Tentativo {attempt} fallito, riprovo...")
    else: 
        print("Errore: impossibile ottenere risultati validi da AI_check_TC")
    
    # Merge LLM results with original input data
    merged_results = {}
    for idx, tc in enumerate(results_list):
        tc_id = tc.get("ID") or tc.get("Id") or tc.get("id")
        #tc_id = tc["ID"]
        if not tc_id:
            print(f" Missing 'ID' field in test case at index {idx}: {tc}")
            continue
        original_data = dic.get(tc_id, {})
 
        merged_tc = {**original_data, **tc}
        merged_results[tc_id] = merged_tc
    
    print(f"Merged result: {merged_results}")
    
    fill_excel_file(merged_results)
    print("File Excel generato con successo!")



if __name__ == "__main__":
    asyncio.run(main())